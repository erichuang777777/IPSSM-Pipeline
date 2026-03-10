#!/usr/bin/env python
"""
IPSSM Pipeline - 一鍵完成資料驗證 + R 風險計算 + 結果整合
合併自 screener_v2.py 和 translator_v3.py

用法:
  python ipssm_pipeline.py data.xlsx          # 一鍵全流程
  python ipssm_pipeline.py data.csv --screen-only          # 僅執行資料驗證
  python ipssm_pipeline.py cleaned.csv --translate-only    # 僅執行 R 計算
  python ipssm_pipeline.py data.xlsx -v validation.xlsx    # 指定驗證檔案比對
"""

import argparse
import csv
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Import cohort converter for handling multiple cohort formats
from cohort_converter import detect_cohort_type, find_column_mapping, STANDARD_IPSSM_COLUMNS

# ============================================================================
#  常數定義
# ============================================================================

# 標準 IPSSM 42 欄
STANDARD_COLUMNS = [
    'ID', 'HB', 'PLT', 'BM_BLAST', 'del5q', 'del7_7q', 'complex', 'CYTO_IPSSR',
    'del17_17p', 'TP53mut', 'TP53maxvaf', 'TP53loh', 'MLL_PTD', 'FLT3', 'ASXL1',
    'BCOR', 'BCORL1', 'CBL', 'CEBPA', 'DNMT3A', 'ETV6', 'EZH2', 'IDH1', 'IDH2',
    'KRAS', 'NF1', 'NPM1', 'NRAS', 'RUNX1', 'SETBP1', 'SF3B1', 'SRSF2', 'STAG2',
    'U2AF1', 'ETNK1', 'GATA2', 'GNB1', 'PHF6', 'PPM1D', 'PRPF8', 'PTPN11', 'WT1'
]

REQUIRED_FIELDS = {'HB', 'PLT', 'BM_BLAST'}

NA_STRINGS = {
    '', ' ', 'NA', 'N/A', 'n/a', 'na', 'NaN', 'nan', 'None', 'none', '.', 'ND', 'nd'
}

VALIDATION_RULES = {
    'HB':         {'min': 4,   'max': 20},
    'PLT':        {'min': 0,   'max': 2000},
    'BM_BLAST':   {'min': 0,   'max': 30},
    'TP53maxvaf': {'min': 0,   'max': 1},
    'CYTO_IPSSR': {'values': ['Very Good', 'Good', 'Intermediate', 'Poor', 'Very Poor']},
    'TP53mut':    {'values': [0, 1, 2]},
}

BINARY_FIELDS = {
    'del5q', 'del7_7q', 'complex', 'del17_17p', 'MLL_PTD', 'FLT3', 'ASXL1',
    'BCOR', 'BCORL1', 'CBL', 'CEBPA', 'DNMT3A', 'ETV6', 'EZH2', 'IDH1', 'IDH2',
    'KRAS', 'NF1', 'NPM1', 'NRAS', 'RUNX1', 'SETBP1', 'SF3B1', 'SRSF2', 'STAG2',
    'U2AF1', 'ETNK1', 'GATA2', 'GNB1', 'PHF6', 'PPM1D', 'PRPF8', 'PTPN11', 'WT1'
}

R_SCRIPT_TEMPLATE = r"""#!/usr/bin/env Rscript

args <- commandArgs(trailingOnly=TRUE)
input_csv    <- args[1]
output_csv   <- args[2]

# Setup library paths (雲端 + 本地 Windows)
lib_candidates <- c(
  path.expand("~/R/library"),
  path.expand("~/AppData/Local/Packages/Claude_pzs8sxrjxfjjc/LocalCache/Local/R/win-library/4.5"),
  path.expand("~/AppData/Local/R/win-library/4.5"),
  "C:/Users/user/AppData/Local/Packages/Claude_pzs8sxrjxfjjc/LocalCache/Local/R/win-library/4.5",
  "C:/Users/user/AppData/Local/R/win-library/4.5"
)

for (lib_path in lib_candidates) {{
  if (dir.exists(lib_path)) .libPaths(c(lib_path, .libPaths()))
}}

suppressPackageStartupMessages({{
  library(ipssm, warn.conflicts=FALSE, verbose=FALSE)
}})

# Read data and check for missing cytogenetic data
data <- read.csv(input_csv, na.strings=c("", " ", "NA", "N/A", "n/a", "na", "NaN", "nan", "None", "none", ".", "ND", "nd"))

# Detect patients with missing cytogenetic fields (scenario analysis triggered)
cytogenetic_fields <- c("del5q", "del7_7q", "del17_17p", "CYTO_IPSSR")
has_missing_cyto <- apply(data[, cytogenetic_fields, drop=FALSE], 1, function(x) any(is.na(x)))

# Run IPSSMwrapper
results <- IPSSMwrapper(input_csv)

# Calculate range (worst_score - best_score) for confidence determination
results$Range_Score <- results$IPSSMscore_worst - results$IPSSMscore_best

# Mark confidence level based on range
results$Confidence_Level <- ifelse(results$Range_Score < 1, "CONFIDENT", "UNCERTAIN")

# Mark if scenario analysis was used (missing cytogenetic data)
results$Used_Scenario_Analysis <- ifelse(has_missing_cyto, "YES", "NO")

# Add detail about which cytogenetic fields were missing
results$Missing_Cytogenetic_Fields <- apply(
  data[, cytogenetic_fields, drop=FALSE],
  1,
  function(x) {{
    missing <- cytogenetic_fields[is.na(x)]
    if (length(missing) > 0) paste(missing, collapse=", ") else "NONE"
  }}
)

# Save results
write.csv(results, file=output_csv, row.names=FALSE, na="NA")

cat("\n[OK] Results generated with scenario analysis and confidence indicators\n")
"""

# ============================================================================
#  第一階段：資料驗證 (Screener)
# ============================================================================

class ValidationReport:
    """收集驗證過程中的錯誤、警告、自動修復和跳過記錄"""

    def __init__(self):
        self.errors = []
        self.warnings = []
        self.auto_fixes = []
        self.skipped_patients = []
        self.input_rows = 0
        self.output_rows = 0
        self.input_cols = 0
        self.output_cols = 0
        self.format_conversions = []

    def add_error(self, row, col, msg):
        self.errors.append(f"Row {row}, {col}: {msg}")

    def add_warning(self, row, col, msg):
        self.warnings.append(f"Row {row}, {col}: {msg}")

    def add_fix(self, row, col, old_val, new_val, reason):
        self.auto_fixes.append(f"Row {row}, {col}: '{old_val}' -> '{new_val}' ({reason})")

    def add_conversion(self, source_col, target_col):
        self.format_conversions.append(f"{source_col} -> {target_col}")

    def skip_patient(self, patient_id, reason):
        self.skipped_patients.append(f"ID={patient_id}: {reason}")

    def report(self):
        lines = [
            f"\nSummary: {len(self.errors)} ERRORS, {len(self.warnings)} WARNINGS, "
            f"{len(self.auto_fixes)} AUTO-FIXES, {len(self.skipped_patients)} SKIPPED\n"
        ]

        if self.format_conversions:
            lines.append("\n--- FORMAT CONVERSIONS ---")
            for conv in self.format_conversions[:10]:
                lines.append(f"  {conv}")
            if len(self.format_conversions) > 10:
                lines.append(f"  ... and {len(self.format_conversions) - 10} more")

        if self.input_rows > 0:
            lines.append("\n--- INFO ---")
            lines.append(f"  [INFO] Input: {self.input_rows} rows x {self.input_cols} columns")
            lines.append(f"  [INFO] Output: {self.output_rows} rows x {self.output_cols} columns (after skipping incomplete records)")
            lines.append(f"  [INFO] Skipped: {len(self.skipped_patients)} patients with missing required fields")

        if self.skipped_patients:
            lines.append(f"\n--- SKIPPED PATIENTS ({len(self.skipped_patients)}) ---")
            for skip in self.skipped_patients[:20]:
                lines.append(f"  [SKIP] {skip}")
            if len(self.skipped_patients) > 20:
                lines.append(f"  ... and {len(self.skipped_patients) - 20} more")

        if self.auto_fixes:
            lines.append(f"\n--- AUTO-FIXES APPLIED ({len(self.auto_fixes)}) ---")
            for fix in self.auto_fixes[:30]:
                lines.append(f"  {fix}")
            if len(self.auto_fixes) > 30:
                lines.append(f"  ... and {len(self.auto_fixes) - 30} more")

        if self.errors:
            lines.append(f"\n--- ERRORS ({len(self.errors)}) ---")
            for error in self.errors[:20]:
                lines.append(f"  {error}")
            if len(self.errors) > 20:
                lines.append(f"  ... and {len(self.errors) - 20} more")

        if self.warnings:
            lines.append(f"\n--- WARNINGS ({len(self.warnings)}) ---")
            for warning in self.warnings[:20]:
                lines.append(f"  {warning}")
            if len(self.warnings) > 20:
                lines.append(f"  ... and {len(self.warnings) - 20} more")

        if len(self.errors) == 0 and len(self.skipped_patients) == 0:
            lines.append("\n>>> PASS: Data is ready for R IPSSMwrapper execution.")
        elif len(self.errors) == 0:
            lines.append(f"\n>>> PASS: {self.output_rows} patients passed validation "
                         f"(after skipping {len(self.skipped_patients)} incomplete records).")
        else:
            lines.append(f"\n>>> FAIL: {len(self.errors)} validation errors found.")

        return '\n'.join(lines) + '\n'


def _read_input_file(input_path):
    """讀取 CSV 或 Excel 檔案，回傳 (rows_list, fieldnames)"""
    input_path = Path(input_path)
    if input_path.suffix.lower() == '.xlsx':
        df = pd.read_excel(input_path, sheet_name=0, dtype=str, keep_default_na=False)
        return df.to_dict('records'), list(df.columns)
    with open(input_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames if reader.fieldnames else []
    return rows, fieldnames


def _try_convert_cohort(input_path, report):
    """
    Attempt to convert input file to standard IPSSM format.
    Returns:
        - (rows, fieldnames, needs_conversion) if conversion applied
        - (rows, fieldnames, False) if already in standard format
    """
    try:
        # Read input file as dataframe
        input_path = Path(input_path)
        if input_path.suffix.lower() == '.xlsx':
            df = pd.read_excel(input_path, sheet_name=0, dtype=str, keep_default_na=False)
        else:
            df = pd.read_csv(input_path, dtype=str, keep_default_na=False)
        
        # Detect cohort type
        cohort_type = detect_cohort_type(df)
        
        # Find column mapping
        mapping = find_column_mapping(df)
        
        # Check if conversion is needed (not all input columns are in standard format)
        is_standard = len(mapping) == len(df.columns) and set(mapping.values()) == set(df.columns)
        
        if cohort_type != "UNKNOWN" and not is_standard:
            print(f"\n  [COHORT DETECTION] Detected format: {cohort_type}")
            print(f"  [COLUMN MAPPING] Found {len(mapping)}/{len(df.columns)} matching columns")
            
            # Apply column mapping
            df_converted = df.rename(columns=mapping)
            
            # Add missing columns with NA
            missing_cols = set(STANDARD_IPSSM_COLUMNS) - set(df_converted.columns)
            if missing_cols:
                print(f"  [MISSING COLUMNS] Adding {len(missing_cols)} missing columns: {', '.join(sorted(missing_cols)[:5])}{'...' if len(missing_cols) > 5 else ''}")
                for col in missing_cols:
                    df_converted[col] = 'NA'
            
            # Reorder to standard format
            df_converted = df_converted[STANDARD_IPSSM_COLUMNS]
            
            # Convert back to rows/fieldnames format
            rows = df_converted.to_dict('records')
            fieldnames = list(df_converted.columns)
            
            report.add_conversion(f"Cohort type '{cohort_type}'", "Standard IPSSM format")
            return rows, fieldnames, True
        
        # Already in standard format or unknown format
        rows, fieldnames = _read_input_file(input_path)
        return rows, fieldnames, False
        
    except Exception as e:
        print(f"  [CONVERSION WARNING] Could not auto-detect cohort format: {e}")
        # Fall back to regular input reading
        rows, fieldnames = _read_input_file(input_path)
        return rows, fieldnames, False


def _convert_fjuh_format(rows, fieldnames, report):
    """偵測並修復欄名尾部空格格式"""
    has_fjuh = any(col.endswith(' ') for col in fieldnames)
    if not has_fjuh:
        return rows

    print("  偵測到異常格式（欄名有尾部空格），自動修復中...")
    mapping = {}
    for col in fieldnames:
        clean_col = col.strip()
        if col != clean_col:
            mapping[col] = clean_col
            report.add_conversion(col, clean_col)

    converted = []
    for row in rows:
        new_row = {}
        for key, value in row.items():
            clean_key = mapping.get(key, key.strip() if isinstance(key, str) else key)
            new_row[clean_key] = value
        converted.append(new_row)
    return converted


def _validate_row(row_idx, row, report):
    """驗證單一資料列，回傳 True=有效 / False=跳過"""
    patient_id = row.get('ID', f'Row{row_idx}')

    # 必填欄位檢查
    missing = [f for f in REQUIRED_FIELDS if row.get(f, '').strip() in NA_STRINGS or row.get(f, '').strip() == '']
    if missing:
        report.skip_patient(patient_id, f"Missing required field(s): {', '.join(missing)}")
        return False

    # NA 標準化
    for col, value in row.items():
        if isinstance(value, str) and value in NA_STRINGS:
            row[col] = 'NA'
            report.add_fix(row_idx, col, value, 'NA', 'Converted NA-like value to standard NA')

    # 數值範圍檢查
    for col in ['HB', 'PLT', 'BM_BLAST', 'TP53maxvaf']:
        value = row.get(col, 'NA')
        if isinstance(value, str):
            value = value.strip()
        if value != 'NA':
            try:
                num = float(value)
                rule = VALIDATION_RULES.get(col)
                if rule and 'min' in rule:
                    if num < rule['min'] or num > rule['max']:
                        report.add_error(row_idx, col, f"Value {num} out of range [{rule['min']}-{rule['max']}]")
            except ValueError:
                report.add_error(row_idx, col, f"Invalid number: {value}")

    # 二元欄位檢查
    for col in BINARY_FIELDS:
        value = row.get(col, 'NA')
        if isinstance(value, str):
            value = value.strip()
        if value != 'NA' and value not in ['0', '1']:
            report.add_error(row_idx, col, f"Binary field must be 0 or 1, got: {value}")

    # 分類欄位檢查
    cyto = row.get('CYTO_IPSSR', 'NA')
    if isinstance(cyto, str):
        cyto = cyto.strip()
    if cyto not in ['NA', 'Very Good', 'Good', 'Intermediate', 'Poor', 'Very Poor']:
        report.add_error(row_idx, 'CYTO_IPSSR', f"Invalid category: {cyto}")

    tp53 = row.get('TP53mut', 'NA')
    if isinstance(tp53, str):
        tp53 = tp53.strip()
    if tp53 not in ['NA', '0', '1', '2', '2 or more']:
        report.add_error(row_idx, 'TP53mut', f"Invalid value: {tp53}")

    return len(report.errors) == 0


def run_screening(input_path, output_path, log_path):
    """執行資料驗證（第一階段），回傳是否成功"""
    print(f"\n{'='*60}")
    print(f"  [階段 1] 資料驗證 & 格式轉換")
    print(f"{'='*60}")
    print(f"  輸入:  {input_path}")
    print(f"  輸出:  {output_path}")
    print(f"  日誌:  {log_path}\n")

    report = ValidationReport()
    valid_rows = []

    try:
        # Try to convert cohort format first
        rows, fieldnames, converted = _try_convert_cohort(input_path, report)
        if converted:
            print(f"  [OK] Successfully converted to standard IPSSM format\n")
        
        report.input_cols = len(fieldnames)
        report.input_rows = len(rows)

        rows = _convert_fjuh_format(rows, fieldnames, report)

        for row_idx, row in enumerate(rows, start=2):
            cleaned = {
                (k.strip() if k else k): (v.strip() if isinstance(v, str) else v)
                for k, v in row.items()
            }
            if _validate_row(row_idx, cleaned, report):
                valid_rows.append(cleaned)

        final_rows = [{col: row.get(col, 'NA') for col in STANDARD_COLUMNS} for row in valid_rows]
        report.output_rows = len(final_rows)
        report.output_cols = len(STANDARD_COLUMNS)

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=STANDARD_COLUMNS)
            writer.writeheader()
            writer.writerows(final_rows)

        report_text = report.report()
        print(report_text)

        with open(log_path, 'w', encoding='utf-8') as f:
            f.write("=" * 70 + "\n")
            f.write("IPSSM Screener Validation Report\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 70 + "\n")
            f.write(report_text)

        print(f"  [OK] Cleaned CSV: {output_path}")
        print(f"  [OK] 驗證日誌:    {log_path}")
        return len(report.errors) == 0

    except Exception as e:
        print(f"  [ERROR] ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
#  第二階段：R 計算 + 結果整合 (Translator)
# ============================================================================

def _find_rscript():
    """自動尋找 Rscript 路徑"""
    candidates = [
        r"C:\Program Files\R\R-4.5.2\bin\Rscript.exe",
        r"C:\Program Files\R\R-4.4.2\bin\Rscript.exe",
        r"C:\Program Files\R\R-4.3.0\bin\Rscript.exe",
    ]
    for c in candidates:
        if Path(c).exists():
            return c
    result = subprocess.run(['where', 'Rscript.exe'], capture_output=True, text=True)
    if result.returncode == 0:
        return result.stdout.strip().split('\n')[0]
    return None


def _read_r_output(csv_path):
    """讀取 R 輸出 CSV"""
    with open(csv_path, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def _read_validation_data(validation_path):
    """從 Excel 讀取驗證數據"""
    try:
        if validation_path and validation_path.exists():
            df = pd.read_excel(validation_path, sheet_name=0, dtype=str)
            result = {}
            for _, row in df.iterrows():
                pid = str(row.get('ID', '')).strip() if 'ID' in row else None
                if pid:
                    result[pid] = dict(row)
            return result
    except Exception as e:
        print(f"  Warning: 無法讀取驗證檔案: {e}")
    return {}


def _save_excel(r_results, output_path, validation_data=None):
    """儲存結果到多工作表 Excel"""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    uncertain_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    uncertain_font = Font(bold=True, color="FFFFFF")

    # --- Sheet 1: Summary ---
    ws = wb.create_sheet("Summary", 0)
    ws.append(['ID', 'Confidence_Level'])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in r_results:
        ws.append([row.get('ID', ''), row.get('Confidence_Level', 'NA')])
        if row.get('Confidence_Level') == 'UNCERTAIN':
            for cell in ws[ws.max_row]:
                cell.fill = uncertain_fill
                cell.font = uncertain_font

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16

    # --- Sheet 2: R_Full_Output ---
    ws_r = wb.create_sheet("R_Full_Output", 1)
    if r_results:
        headers = list(r_results[0].keys())
        ws_r.append(headers)
        for cell in ws_r[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in r_results:
            ws_r.append([row.get(h, '') for h in headers])
        for col in ws_r.columns:
            ws_r.column_dimensions[col[0].column_letter].width = 14

    # --- Sheet 3: Analysis (if validation data available) ---
    if validation_data:
        ws_a = wb.create_sheet("Analysis", 2)
        a_headers = ['ID', 'R_Confidence', 'R_Category', 'Validation_Result', 'Match', 'Range_Score', 'Notes']
        ws_a.append(a_headers)
        for cell in ws_a[1]:
            cell.fill = header_fill
            cell.font = header_font

        match_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in r_results:
            pid = row.get('ID', '')
            r_conf = row.get('Confidence_Level', 'NA')
            r_cat = row.get('IPSSMcat', 'NA')
            r_range = row.get('Range_Score', 'NA')
            val = validation_data.get(pid, {})
            val_result = val.get('IPSS_M_', 'N/A') if val else 'N/A'
            match = 'YES' if r_cat == val_result else ('NO' if val_result != 'N/A' else 'N/A')
            notes = ''
            if match == 'NO':
                notes = f"Expected: {val_result}, Got: {r_cat}"
            elif r_conf == 'UNCERTAIN':
                notes = "Wide range, check carefully"

            ws_a.append([pid, r_conf, r_cat, val_result, match, r_range, notes])
            cur = ws_a.max_row
            if match == 'YES':
                for cell in ws_a[cur]:
                    cell.fill = match_fill
            elif match == 'NO':
                for cell in ws_a[cur]:
                    cell.fill = mismatch_fill

        for letter, width in zip('ABCDEFG', [12, 14, 14, 16, 10, 12, 30]):
            ws_a.column_dimensions[letter].width = width

    wb.save(output_path)


def run_translation(input_csv, rscript_path=None, validation_path=None):
    """執行 R 計算 + 結果整合（第二階段），回傳是否成功"""
    print(f"\n{'='*60}")
    print(f"  [階段 2] R 風險計算 & 信心等級標記")
    print(f"{'='*60}")

    input_path = Path(input_csv)
    rscript = rscript_path or _find_rscript()
    if not rscript:
        print("  [ERROR] ERROR: 找不到 Rscript！請安裝 R >= 4.3.0")
        return False

    output_dir = input_path.parent
    r_output_csv = output_dir / f"{input_path.stem}_r_output.csv"
    excel_output = output_dir / f"{input_path.stem}_results.xlsx"

    # 尋找驗證檔案
    if not validation_path:
        potential = output_dir / "IPSSM_validation_result.xlsx"
        if potential.exists():
            validation_path = potential

    print(f"  輸入:    {input_path}")
    print(f"  Rscript: {rscript}")
    print(f"  輸出:    {excel_output}")
    if validation_path:
        print(f"  驗證檔:  {validation_path}")

    # 建立並執行 R 腳本
    with tempfile.NamedTemporaryFile(mode='w', suffix='.R', delete=False) as f:
        f.write(R_SCRIPT_TEMPLATE)
        r_script_file = f.name

    try:
        print("\n  執行 R IPSSMwrapper...")
        cmd = [rscript, r_script_file, str(input_path), str(r_output_csv)]
        result = subprocess.run(cmd, capture_output=True, text=True)

        if result.stdout:
            print(result.stdout)
        if result.stderr:
            print("  [STDERR]:", result.stderr)

        if result.returncode != 0:
            print("  [ERROR] ERROR: R 執行失敗")
            with open(output_dir / "r_error.log", "w", encoding="utf-8") as f:
                f.write(result.stderr or "No stderr output recorded.")
            return False

        if not r_output_csv.exists():
            print("  [ERROR] ERROR: R 輸出檔案未產生")
            return False

        # 處理結果
        print("  處理結果中...")
        r_results = _read_r_output(r_output_csv)

        validation_data = {}
        if validation_path:
            validation_data = _read_validation_data(validation_path)

        confident = sum(1 for r in r_results if r.get('Confidence_Level') == 'CONFIDENT')
        uncertain = sum(1 for r in r_results if r.get('Confidence_Level') == 'UNCERTAIN')

        print(f"\n  === 信心等級統計 ===")
        print(f"  CONFIDENT (range < 1):  {confident}")
        print(f"  UNCERTAIN (range >= 1): {uncertain}")
        print(f"  總計: {len(r_results)}\n")

        _save_excel(r_results, excel_output, validation_data)

        print(f"  [OK] 結果已儲存: {excel_output}")
        print(f"    Sheet 'Summary':       ID + Confidence_Level")
        print(f"    Sheet 'R_Full_Output':  完整 R 計算數據")
        if validation_data:
            print(f"    Sheet 'Analysis':      與驗證數據比對")
        return True

    finally:
        Path(r_script_file).unlink(missing_ok=True)


# ============================================================================
#  主程式入口
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='IPSSM Pipeline - 一鍵完成資料驗證 + R 風險計算',
        epilog='支援 CSV 和 Excel 輸入，自動修復異常空格格式。'
    )
    parser.add_argument('input_file', help='輸入檔案（CSV 或 Excel）')
    parser.add_argument('-v', '--validation', help='手動驗證結果 Excel 檔案（比對用）')
    parser.add_argument('--rscript', help='Rscript 執行路徑')
    parser.add_argument('--screen-only', action='store_true', help='僅執行資料驗證，不執行 R 計算')
    parser.add_argument('--translate-only', action='store_true', help='僅執行 R 計算（輸入須為已清理的 CSV）')

    args = parser.parse_args()
    input_path = Path(args.input_file)

    if not input_path.exists():
        print(f"ERROR: 輸入檔案不存在: {input_path}")
        sys.exit(1)

    validation_path = Path(args.validation) if args.validation else None

    print(f"\n{'#'*60}")
    print(f"  IPSSM Pipeline")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'#'*60}")

    # === 僅翻譯模式 ===
    if args.translate_only:
        success = run_translation(str(input_path), args.rscript, validation_path)
        print(f"\n{'='*60}")
        print(f"  完成！{'[OK] 成功' if success else '[ERROR] 失敗'}")
        print(f"{'='*60}")
        sys.exit(0 if success else 1)

    # === 執行驗證 ===
    cleaned_csv = input_path.parent / f"{input_path.stem}_cleaned.csv"
    log_path = input_path.parent / f"{input_path.stem}_screening_log.txt"

    screen_ok = run_screening(input_path, cleaned_csv, log_path)

    if not screen_ok:
        print("\n  [ERROR] 驗證階段有錯誤，請檢查日誌。")
        sys.exit(1)

    # === 僅驗證模式 ===
    if args.screen_only:
        print(f"\n{'='*60}")
        print(f"  完成！（僅驗證模式）")
        print(f"{'='*60}")
        sys.exit(0)

    # === 執行 R 計算 ===
    translate_ok = run_translation(str(cleaned_csv), args.rscript, validation_path)

    print(f"\n{'='*60}")
    print(f"  Pipeline 完成！{'[OK] 全部成功' if translate_ok else '[ERROR] R 計算失敗'}")
    print(f"{'='*60}")
    sys.exit(0 if translate_ok else 1)


if __name__ == '__main__':
    main()
